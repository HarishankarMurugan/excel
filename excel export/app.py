from flask import Flask, render_template, request, send_file
import openpyxl
import io

app = Flask(__name__)

# Function to calculate y1, y2, y3, y4 based on given x1, x2, x3, x4, a, b
def calculate_values(x1, x2, x3, x4, a, b):
    y1 = (x1 + x2) * (x3 - x4) + (a / b)
    y2 = (x1 * x2) + (x3 / x4) - a
    y3 = (x1 * x2) * (x3 + x4)
    y4 = (x1 / x2) + (x3 * x4) - (a + b)
    return y1, y2, y3, y4

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/download', methods=['POST'])
def download():
    x1 = float(request.form['x1'])
    x2 = float(request.form['x2'])
    x3 = float(request.form['x3'])
    x4 = float(request.form['x4'])
    a_max = int(request.form['a_max'])
    b_max = int(request.form['b_max'])

    # Create a new Workbook and select the active sheet
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Output"

    # Fill column headers (1 in B1, 2 in D1, 3 in F1, ...)
    for col in range(1, a_max + 1):
        sheet.cell(row=1, column=2 * col).value = f"{col}"

    # Fill row headers (1 in A2, 2 in A4, 3 in A6, ...)
    for row in range(1, b_max + 1):
        sheet.cell(row=2 * row, column=1).value = f"{row}"

    # Populate the Excel sheet with calculated values
    for row in range(1, b_max + 1):
        for col in range(1, a_max + 1):
            current_a = col
            current_b = row
            y1, y2, y3, y4 = calculate_values(x1, x2, x3, x4, current_a, current_b)

            sheet.cell(row=2 * row, column=2 * col).value = y1
            sheet.cell(row=2 * row, column=2 * col + 1).value = y2
            sheet.cell(row=2 * row + 1, column=2 * col).value = y3
            sheet.cell(row=2 * row + 1, column=2 * col + 1).value = y4

    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name="Workout result.xlsx", as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
